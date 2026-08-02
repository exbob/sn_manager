from datetime import date
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook

from sn_manager.app.export import (
    export_burn_and_mark_used,
    export_burn_txt,
    export_excel,
    export_selected_and_mark_used,
)
from sn_manager.app.services import SnService
from sn_manager.core.status import Status
from sn_manager.db import master_data as md
from sn_manager.db.connection import connect


def _seed_svg14(conn) -> None:
    md.upsert_product(conn, "SVG14", "示例外壳机")
    md.upsert_hardware_batch(conn, "SVG14", "05", "第五批")


def test_excel_and_burn(tmp_path: Path):
    rows = [
        {
            "sn": "ASVG140521261CF04",
            "product_model": "SVG14",
            "hw_batch": "05",
            "factory": "2",
            "market": "1",
            "prod_year": 2026,
            "prod_month": 1,
            "prod_day": 12,
            "seq": 0xF04,
            "status": "unused",
            "created_at": "t",
            "updated_at": "t",
        }
    ]
    xlsx = tmp_path / "out.xlsx"
    export_excel(rows, xlsx)
    wb = load_workbook(xlsx)
    assert wb.active["A2"].value == "ASVG140521261CF04"

    paths = export_burn_txt(["ASVG140521261CF04"], tmp_path)
    p = tmp_path / "sn_ASVG140521261CF04.txt"
    assert p in paths
    assert p.read_text(encoding="utf-8") == "ASVG140521261CF04"


def test_burn_txt_overwrites_existing(tmp_path: Path):
    sn = "ASVG140521261CF04"
    p = tmp_path / f"sn_{sn}.txt"
    p.write_text("old", encoding="utf-8")

    export_burn_txt([sn], tmp_path)

    assert p.read_text(encoding="utf-8") == sn


def test_export_burn_and_mark_used(tmp_path: Path):
    conn = connect(tmp_path / "t.db")
    _seed_svg14(conn)
    svc = SnService(conn)
    rows = svc.generate(
        product_model="SVG14",
        hw_batch="05",
        factory="2",
        market="1",
        prod_date=date(2026, 1, 12),
        count=1,
    )
    sn = rows[0]["sn"]
    burn_dir = tmp_path / "burn"

    export_burn_and_mark_used(svc, [sn], burn_dir, mark_used=True)

    assert (burn_dir / f"sn_{sn}.txt").read_text(encoding="utf-8") == sn
    assert svc.filter(sn=sn)[0]["status"] == Status.USED.value


def test_export_burn_failure_does_not_mark_used(tmp_path: Path, monkeypatch):
    svc = MagicMock()
    svc.set_status = MagicMock()

    def boom(*_args, **_kwargs):
        raise OSError("write failed")

    monkeypatch.setattr("sn_manager.app.export.export_burn_txt", boom)

    with pytest.raises(OSError):
        export_burn_and_mark_used(
            svc,
            ["ASVG140521261CF04"],
            tmp_path,
            mark_used=True,
        )

    svc.set_status.assert_not_called()


def test_export_selected_both_mark_used(tmp_path: Path):
    conn = connect(tmp_path / "t.db")
    _seed_svg14(conn)
    svc = SnService(conn)
    rows = svc.generate(
        product_model="SVG14",
        hw_batch="05",
        factory="2",
        market="1",
        prod_date=date(2026, 1, 12),
        count=1,
    )
    sn = rows[0]["sn"]
    xlsx = tmp_path / "20260731152950.xlsx"

    export_selected_and_mark_used(
        svc,
        rows,
        burn=True,
        excel=True,
        export_directory=tmp_path,
        mark_used=True,
        excel_path=xlsx,
    )

    assert (tmp_path / f"sn_{sn}.txt").read_text(encoding="utf-8") == sn
    assert load_workbook(xlsx).active["A2"].value == sn
    assert svc.filter(sn=sn)[0]["status"] == Status.USED.value


def test_export_selected_excel_failure_does_not_mark_used(tmp_path: Path, monkeypatch):
    conn = connect(tmp_path / "t.db")
    _seed_svg14(conn)
    svc = SnService(conn)
    rows = svc.generate(
        product_model="SVG14",
        hw_batch="05",
        factory="2",
        market="1",
        prod_date=date(2026, 1, 12),
        count=1,
    )
    sn = rows[0]["sn"]

    def boom(*_a, **_k):
        raise OSError("excel failed")

    monkeypatch.setattr("sn_manager.app.export.export_excel", boom)

    with pytest.raises(OSError):
        export_selected_and_mark_used(
            svc,
            rows,
            burn=True,
            excel=True,
            export_directory=tmp_path,
            mark_used=True,
            excel_path=tmp_path / "out.xlsx",
        )

    assert (tmp_path / f"sn_{sn}.txt").exists()
    assert svc.filter(sn=sn)[0]["status"] == Status.UNUSED.value
